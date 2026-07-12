#!/usr/bin/env python3
"""
build_site.py -- generate the static dashboard site (_site/) from Pool_Log.xlsx,
photos/, and the testing guide. Run after append_log.py/format_log.py so the
site reflects the latest logged data.

Output (_site/, gitignored -- rebuilt fresh every run, never committed):
  index.html              from site_src/, with style.css's href version-busted
                          (?v=<build timestamp>) so a redeploy never pairs new
                          HTML against a visitor's stale cached stylesheet
  style.css               copied verbatim from site_src/
  data.json               TEST/DOSE rows + season meta, for index.html's charts
                          (fetched with cache: 'no-store' -- always fresh)
  log.html                full raw log table, generated directly
  photos.html             day-by-day photo gallery, generated directly
  photos/*.jpg            resized copies of photos/ (max 1600px, ~q82)
  Pool_Log.xlsx           copied as-is, for the download link
  Taylor_K2006_Testing_Guide.docx   copied as-is, for the download link

Usage:  python build_site.py
"""

import json
import os
import shutil
from datetime import datetime, date

from openpyxl import load_workbook
from PIL import Image, ImageOps

from append_log import COLS as C, LOG_FILE, SHEET
import dose as D

HERE = os.path.dirname(os.path.abspath(__file__))
OUT = os.path.join(HERE, "_site")
PHOTOS_SRC = os.path.join(HERE, "photos")
CONCERT_SRC = os.path.join(HERE, "concert")
CONCERT2_SRC = os.path.join(HERE, "concert2")
GUIDE_SRC = os.path.join(HERE, "Taylor_K2006_Testing_Guide.docx")

MAX_PHOTO_DIM = 1600
JPEG_QUALITY = 82


def cell(ws, r, key):
    return ws.cell(r, C[key]).value


def load_rows():
    wb = load_workbook(LOG_FILE, data_only=False)
    ws = wb[SHEET]
    tests, doses, all_rows = [], [], []
    for r in range(3, ws.max_row + 1):
        typ = cell(ws, r, "type")
        if not typ:
            continue
        row = {k: cell(ws, r, k) for k in C}
        all_rows.append(row)
        if typ == "TEST" and isinstance(row["fc"], (int, float)):
            tests.append(row)
        elif typ == "DOSE" and isinstance(row["chlorine_gal"], (int, float)):
            doses.append(row)
    return tests, doses, all_rows


def build_data_json(tests, doses, all_rows, photo_by_date):
    cfg = D.CONFIG
    floor, aim = D.band_for_cya(cfg["cya_current"], cfg)
    season_total = round(sum(d["chlorine_gal"] for d in doses), 2)
    open_row = next((r for r in all_rows if r["type"] and "Open" in str(r["type"])), None)
    season_open = str(open_row["date"]) if open_row else (str(tests[0]["date"]) if tests else None)
    data = {
        "meta": {
            "generated_at": datetime.now().isoformat(timespec="seconds"),
            "pool_gallons": cfg["pool_gallons"],
            "cya_current": cfg["cya_current"],
            "fc_floor": floor,
            "fc_aim": aim,
            "season_open": season_open,
            "season_total_dose_gal": season_total,
        },
        "tests": [
            {
                "date": str(t["date"]), "time": t["time"], "fc": t["fc"], "cc": t["cc"],
                "water_temp": t["water_temp"], "sun_hrs": t["sun_hrs"],
                "pred_error": t["pred_error"],
                "fc_loss": t["fc_loss"], "ph": t["ph"], "ta": t["ta"], "cya": t["cya"],
                "rain_in": t["rain_in"], "water_level_cm": t["water_level_cm"],
                "thumb": (photo_by_date.get(str(t["date"])) or [None])[0],
            }
            for t in tests
        ],
        "doses": [
            {
                "date": str(d["date"]), "time": d["time"], "gal": d["chlorine_gal"],
                "cum": d["cum_cl"], "pred_fc": d["pred_fc"],
            }
            for d in doses
        ],
        "fills": [
            {"date": str(r["date"]), "gal": r["fill_gal"]}
            for r in all_rows
            if r["type"] and "Fill" in str(r["type"]) and isinstance(r["fill_gal"], (int, float))
        ],
    }
    with open(os.path.join(OUT, "data.json"), "w") as fh:
        json.dump(data, fh, indent=2, default=str)
    return data


LOG_COLUMNS = [
    ("date", "Date"), ("time", "Time"), ("type", "Type"), ("ph", "pH"),
    ("fc", "FC"), ("cc", "CC"), ("fc_loss", "FC loss"), ("pred_fc", "Pred FC"),
    ("pred_error", "Pred err"), ("ta", "TA"), ("ch", "CH"), ("cya", "CYA"),
    ("chlorine_gal", "Cl added"), ("cum_cl", "Cum Cl"), ("sun_hrs", "Sun"),
    ("rain_in", "Rain"), ("fill_gal", "Fill"), ("water_temp", "Water temp"),
    ("water_level_cm", "Water level"),
    ("weather", "Weather"), ("photo", "Photo"), ("notes", "Notes"),
]


def esc(v):
    if v is None:
        return ""
    return str(v).replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;")


def build_log_html(all_rows, photo_days, build_version):
    def col_sum(key, places):
        return round(sum(r[key] for r in all_rows if isinstance(r[key], (int, float))), places)

    totals = {"chlorine_gal": col_sum("chlorine_gal", 2), "rain_in": col_sum("rain_in", 2), "fill_gal": col_sum("fill_gal", 1)}

    head = "".join(f"<th>{label}</th>" for _, label in LOG_COLUMNS)

    totals_cells = ['<td colspan="4">SEASON TOTALS</td>']
    for i, (key, _) in enumerate(LOG_COLUMNS):
        if i < 4:
            continue
        v = totals.get(key)
        totals_cells.append(f"<td>{v if v is not None else ''}</td>")
    totals_row = f'<tr class="totals-row">{"".join(totals_cells)}</tr>'

    body_rows = []
    prev_date, shade = None, True  # first distinct date ends up unshaded (matches the Excel log)
    for row in all_rows:
        d = str(row["date"])
        if d != prev_date:
            shade = not shade
            prev_date = d
        classes = ["row-" + row["type"].split(" ")[0].lower()]
        if shade:
            classes.append("day-alt")
        cells = []
        for key, _ in LOG_COLUMNS:
            val = esc(row[key])
            if key == "date" and d in photo_days:
                cells.append(f'<td><a href="photos.html#{d}">{val}</a></td>')
            else:
                cells.append(f"<td>{val}</td>")
        body_rows.append(f'<tr class="{" ".join(classes)}">{"".join(cells)}</tr>')
    html = f"""<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="utf-8"><meta name="viewport" content="width=device-width, initial-scale=1">
<script src="theme.js?v={build_version}"></script>
<title>Full log — Summer 2026 Pool Data</title>
<link rel="preconnect" href="https://fonts.googleapis.com">
<link rel="stylesheet" href="https://fonts.googleapis.com/css2?family=DM+Sans:wght@400;500&family=Fraunces:wght@500&display=swap">
<link rel="stylesheet" href="style.css?v={build_version}">
</head>
<body>
<header class="masthead">
<a class="wordmark" href="index.html">Summer 2026 Pool Data</a>
<div style="display:flex;align-items:center;gap:20px;flex-wrap:wrap;justify-content:center;">
<nav><a href="index.html">Dashboard</a><a href="log.html" class="active">Data</a><a href="photos.html">Photos</a><a href="Pool_Log.xlsx">Download Full Log</a><a href="Taylor_K2006_Testing_Guide.docx">How I test</a></nav>
<button type="button" id="themeToggle" class="theme-toggle" aria-label="Toggle light/dark theme"></button>
</div>
</header>
<script>
(function () {{
  var btn = document.getElementById('themeToggle');
  function labelToggle() {{
    var next = window.getTheme() === 'dark' ? 'light' : 'dark';
    btn.innerHTML = window.THEME_ICONS[next === 'dark' ? 'moon' : 'sun'];
    btn.setAttribute('aria-label', 'Switch to ' + next + ' mode');
  }}
  labelToggle();
  btn.addEventListener('click', function () {{
    var next = window.getTheme() === 'dark' ? 'light' : 'dark';
    localStorage.setItem('pool_theme', next);
    document.documentElement.setAttribute('data-theme', next);
    location.reload();
  }});
}})();
</script>
<main class="full">
<h1>Full season log</h1>
<div class="table-wrap">
<table class="logtable">
<thead>
<tr>{head}</tr>
{totals_row}
</thead>
<tbody>
{''.join(body_rows)}
</tbody>
</table>
</div>
</main>
<script>
(function () {{
  const wrap = document.querySelector('.table-wrap');
  const headTr = wrap.querySelector('thead tr:first-child');
  const totalsTr = wrap.querySelector('tr.totals-row');
  if (headTr && totalsTr) {{
    const h = headTr.getBoundingClientRect().height;
    totalsTr.querySelectorAll('td').forEach(td => {{ td.style.top = h + 'px'; }});
  }}
  let isDown = false, startX = 0, startScroll = 0;
  wrap.addEventListener('mousedown', e => {{
    isDown = true; wrap.classList.add('dragging');
    startX = e.pageX; startScroll = wrap.scrollLeft;
  }});
  window.addEventListener('mouseup', () => {{ isDown = false; wrap.classList.remove('dragging'); }});
  window.addEventListener('mousemove', e => {{
    if (!isDown) return;
    e.preventDefault();
    wrap.scrollLeft = startScroll - (e.pageX - startX);
  }});
}})();
</script>
</body>
</html>"""
    with open(os.path.join(OUT, "log.html"), "w", encoding="utf-8") as fh:
        fh.write(html)


def resize_photo(src_path, dst_path):
    with Image.open(src_path) as im:
        # Phone photos often carry an EXIF orientation tag (raw sensor data is
        # landscape, tag says "rotate N to display upright") -- most viewers
        # respect it automatically, but a naive resize+resave silently drops
        # it, leaving the web copy sideways even though the original is fine.
        im = ImageOps.exif_transpose(im)
        im = im.convert("RGB")
        w, h = im.size
        if max(w, h) > MAX_PHOTO_DIM:
            scale = MAX_PHOTO_DIM / max(w, h)
            im = im.resize((round(w * scale), round(h * scale)), Image.LANCZOS)
        im.save(dst_path, "JPEG", quality=JPEG_QUALITY, optimize=True)


def scan_and_resize_photos():
    # Photos are resized here; videos are expected to already be web-ready,
    # compressed .mp4 (transcoded from the phone's .MOV with ffmpeg before
    # committing -- see POOL.md; the large .MOV originals are gitignored, only
    # the small .mp4 lands in the repo). We just copy the .mp4 through.
    out_photos = os.path.join(OUT, "photos")
    os.makedirs(out_photos, exist_ok=True)
    by_date, videos_by_date = {}, {}
    if os.path.isdir(PHOTOS_SRC):
        for fname in sorted(os.listdir(PHOTOS_SRC)):
            low = fname.lower()
            day = fname.split("_")[0]
            src = os.path.join(PHOTOS_SRC, fname)
            if low.endswith((".jpg", ".jpeg", ".png")):
                resize_photo(src, os.path.join(out_photos, fname))
                by_date.setdefault(day, []).append(fname)
            elif low.endswith(".mp4"):
                shutil.copy(src, os.path.join(out_photos, fname))
                videos_by_date.setdefault(day, []).append(fname)
    return by_date, videos_by_date


def build_photos(tests, by_date, videos_by_date, build_version):
    floor, _ = D.band_for_cya(D.CONFIG["cya_current"], D.CONFIG)
    test_by_date = {str(t["date"]): t for t in tests}

    def stat_tile(label, value):
        return f'<div><div class="stat-label">{label}</div><div class="stat-value">{esc(value)}</div></div>'

    def day_stats_html(day):
        t = test_by_date.get(day)
        if not t or not isinstance(t.get("fc"), (int, float)):
            return ""
        in_band = t["fc"] >= floor
        tiles = [
            f'<div><div class="stat-label">Free chlorine</div><div class="stat-value big serif" style="color:var(--teal);">{t["fc"]:.1f}</div></div>',
            f'<span class="pill {"good" if in_band else "watch"}">{"in target band" if in_band else "below floor"}</span>',
            stat_tile("Combined chlorine", f'{t["cc"]:.1f}' if isinstance(t.get("cc"), (int, float)) else "—"),
        ]
        if isinstance(t.get("water_temp"), (int, float)):
            tiles.append(stat_tile("Water temp", f'{t["water_temp"]:.1f}°F'))
        if isinstance(t.get("sun_hrs"), (int, float)):
            tiles.append(stat_tile("Sun that day", f'{t["sun_hrs"]:.1f} hrs'))
        return f'<div style="display:flex;align-items:center;justify-content:center;gap:20px;flex-wrap:wrap;margin-top:12px;">{"".join(tiles)}</div>'

    sections = []
    # Include days that have a video even if they have no photos.
    all_days = sorted(set(by_date) | set(videos_by_date), reverse=True)
    for day in all_days:
        videos = "".join(
            f'<video class="day-video" controls preload="metadata" playsinline>'
            f'<source src="photos/{v}" type="video/mp4">Your browser can\'t play this video.</video>'
            for v in videos_by_date.get(day, [])
        )
        thumbs = "".join(
            f'<img src="photos/{f}" loading="lazy" alt="Chlorine-check photo, {day}">'
            for f in by_date.get(day, [])
        )
        sections.append(f'<section class="photo-day" id="{day}"><h2>{day}</h2>{videos}<div class="thumbs">{thumbs}</div>{day_stats_html(day)}</section>')

    html = f"""<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="utf-8"><meta name="viewport" content="width=device-width, initial-scale=1">
<script src="theme.js?v={build_version}"></script>
<title>Photos — Summer 2026 Pool Data</title>
<link rel="preconnect" href="https://fonts.googleapis.com">
<link rel="stylesheet" href="https://fonts.googleapis.com/css2?family=DM+Sans:wght@400;500&family=Fraunces:wght@500&display=swap">
<link rel="stylesheet" href="style.css?v={build_version}">
</head>
<body>
<header class="masthead">
<a class="wordmark" href="index.html">Summer 2026 Pool Data</a>
<div style="display:flex;align-items:center;gap:20px;flex-wrap:wrap;justify-content:center;">
<nav><a href="index.html">Dashboard</a><a href="log.html">Data</a><a href="photos.html" class="active">Photos</a><a href="Pool_Log.xlsx">Download Full Log</a><a href="Taylor_K2006_Testing_Guide.docx">How I test</a></nav>
<button type="button" id="themeToggle" class="theme-toggle" aria-label="Toggle light/dark theme"></button>
</div>
</header>
<script>
(function () {{
  var btn = document.getElementById('themeToggle');
  function labelToggle() {{
    var next = window.getTheme() === 'dark' ? 'light' : 'dark';
    btn.innerHTML = window.THEME_ICONS[next === 'dark' ? 'moon' : 'sun'];
    btn.setAttribute('aria-label', 'Switch to ' + next + ' mode');
  }}
  labelToggle();
  btn.addEventListener('click', function () {{
    var next = window.getTheme() === 'dark' ? 'light' : 'dark';
    localStorage.setItem('pool_theme', next);
    document.documentElement.setAttribute('data-theme', next);
    location.reload();
  }});
}})();
</script>
<main class="wide">
<h1>Chlorine-check photos</h1>
<p class="cap">Daily clarity/stain-tracking photos, most recent first. Click a photo to zoom, click again to shrink.</p>
{''.join(sections)}
</main>
<div class="lightbox" id="lightbox"><img id="lightboxImg" src="" alt=""></div>
<script>
document.querySelectorAll('.thumbs img').forEach(img => {{
  img.addEventListener('click', () => {{
    document.getElementById('lightboxImg').src = img.src;
    document.getElementById('lightbox').classList.add('open');
  }});
}});
document.getElementById('lightbox').addEventListener('click', () => {{
  document.getElementById('lightbox').classList.remove('open');
}});
</script>
</body>
</html>"""
    with open(os.path.join(OUT, "photos.html"), "w", encoding="utf-8") as fh:
        fh.write(html)


# The two-night Noah Kahan Fenway run. Each night is a standalone shareable gallery
# page. setlist/encore sourced from setlist.fm (fan-submitted -- verify). Web-ready
# media lives in the committed dir; raw camera originals are gitignored (icloud/,
# icloud/night2/) and processed offline by prep_concert*.py.
CONCERTS = [
    {
        "page": "concert.html", "dir": CONCERT_SRC, "web": "concert", "label": "Night 1",
        "heading": "Noah Kahan — July 8, 2026 · Fenway Park, Boston, MA",
        "title": "Noah Kahan — Night 1 (July 8, 2026)",
        "acts": [
            {"name": "Act 1 · Main Stage", "songs": ["American Cars", "Doors", "All My Love", "Deny Deny Deny", "Staying Still"]},
            {"name": "Act 2 · B-Stage", "songs": ["Haircut", "Downfall", "Mess"]},
            {"name": "Act 3 · Main Stage", "songs": [
                {"t": "She Calls Me Back", "note": "with Gigi Perez · Big Papi cameo on the screen before the song"},
                "Dashboard", "Dial Drunk"]},
            {"name": "Act 4 · Roof / Main Stage", "songs": ["We Go Way Back", "Porch Light"]},
            {"name": "Act 5 · C-Stage", "songs": [
                {"t": "Orbiter", "note": "extended outro"}, "Maine", "Paid Time Off",
                {"t": "Lighthouse", "note": "live debut"},
                {"t": "The View Between Villages", "note": "extended"}]},
            {"name": "Act 6 · Main Stage", "songs": ["Northern Attitude", "The Great Divide", "Orange Juice", "New Perspective"]},
            {"name": "Encore", "encore": True, "songs": [
                "End of August", "Homesick",
                {"t": "Stick Season", "note": "extended · fireworks at the end"}]},
        ],
    },
    {
        "page": "concert2.html", "dir": CONCERT2_SRC, "web": "concert2", "label": "Night 2",
        "heading": "Noah Kahan — July 11, 2026 · Fenway Park, Boston, MA",
        "title": "Noah Kahan — Night 2 (July 11, 2026)",
        "acts": [
            {"name": "Act 1 · Main Stage", "songs": ["American Cars", "Doors", "All My Love", "Deny Deny Deny", "Staying Still"]},
            {"name": "Act 2 · B-Stage", "songs": ["Haircut", "Downfall",
                {"t": "Forever", "note": "dedicated to Zuza"}]},
            {"name": "Act 3 · Main Stage", "songs": [
                {"t": "She Calls Me Back", "note": "with Gigi Perez · Noah's mom on the screen reveals he's added to the Fenway Music Hall of Fame"},
                "Dashboard", "Dial Drunk"]},
            {"name": "Act 4 · Roof / Main Stage", "songs": ["Willing and Able", "Porch Light"]},
            {"name": "Act 5 · C-Stage", "songs": ["Orbiter",
                {"t": "23", "note": "live debut"}, "Paid Time Off", "Dan",
                {"t": "The View Between Villages", "note": "extended"}]},
            {"name": "Act 6 · Main Stage", "songs": ["Northern Attitude", "The Great Divide", "Carlo's Song", "New Perspective"]},
            {"name": "Encore", "encore": True, "songs": [
                "End of August", "Homesick",
                {"t": "Stick Season", "note": "extended · fireworks at the end"}]},
        ],
    },
]


def scan_concert_dir(src_dir, web_name):
    """Copy a night's web-ready media (pre-resized photos + already-transcoded H.264
    mp4s) into _site/<web_name>/. Returns (images, videos) filename lists sorted by name
    (iPhone IMG_#### numbering is chronological). Heavy resize/transcode is done once
    offline (prep_concert*.py; see POOL.md), so this -- and the CI build, no ffmpeg -- only copy."""
    if not os.path.isdir(src_dir):
        return [], []
    out = os.path.join(OUT, web_name)
    os.makedirs(out, exist_ok=True)
    images, videos = [], []
    for fname in sorted(os.listdir(src_dir)):
        low = fname.lower()
        if low.endswith((".jpg", ".jpeg", ".png")):
            shutil.copy(os.path.join(src_dir, fname), os.path.join(out, fname))
            images.append(fname)
        elif low.endswith(".mp4"):
            shutil.copy(os.path.join(src_dir, fname), os.path.join(out, fname))
            videos.append(fname)
    return images, videos


def setlist_html(concert):
    """Render the setlist grouped into acts/stages, continuously numbered across acts,
    with optional per-song notes (guest, live debut, extended, cameos). A song is either
    a plain title string or a {"t": title, "note": ...} dict."""
    acts = concert.get("acts")
    if not acts:
        return ""
    blocks, n = [], 1
    for act in acts:
        songs = act["songs"]
        items = []
        for s in songs:
            if isinstance(s, dict):
                note = f'<span class="song-note">{esc(s["note"])}</span>' if s.get("note") else ""
                items.append(f'<li>{esc(s["t"])}{note}</li>')
            else:
                items.append(f"<li>{esc(s)}</li>")
        label_cls = "setlist-act-label encore" if act.get("encore") else "setlist-act-label"
        blocks.append(f'<div class="setlist-act"><div class="{label_cls}">{esc(act["name"])}</div>'
                      f'<ol class="setlist-songs" start="{n}">{"".join(items)}</ol></div>')
        n += len(songs)
    return ('<section class="setlist"><h2 class="concert-sub">Setlist</h2>'
            f'<div class="setlist-acts">{"".join(blocks)}</div></section>')


def _video_dims(path, _cache={}):
    """(width, height) display dims of an mp4's video track from its tkhd box.
    Pure-python (no ffprobe) so it also works on the CI build runner; relies on the
    moov atom being near the front, which it is (all clips are encoded +faststart).
    Used to lay portrait clips out solo/centered and landscape clips two-up."""
    if path in _cache:
        return _cache[path]
    try:
        with open(path, "rb") as f:
            buf = f.read(1 << 20)
    except OSError:
        buf = b""

    def walk(start, end):
        i = start
        while i + 8 <= end:
            size = int.from_bytes(buf[i:i + 4], "big")
            btype = buf[i + 4:i + 8]
            hdr = 8
            if size == 1:
                size = int.from_bytes(buf[i + 8:i + 16], "big"); hdr = 16
            elif size == 0:
                size = end - i
            box_end = min(i + size, end)
            if size < hdr:
                break
            if btype in (b"moov", b"trak", b"mdia"):
                r = walk(i + hdr, box_end)
                if r:
                    return r
            elif btype == b"tkhd":
                w = int.from_bytes(buf[box_end - 8:box_end - 4], "big") >> 16
                h = int.from_bytes(buf[box_end - 4:box_end], "big") >> 16
                if w and h:
                    return (w, h)
            i = box_end
        return None

    _cache[path] = walk(0, len(buf))
    return _cache[path]


def build_concert_page(concert, all_concerts, build_version):
    """One standalone shareable gallery page per concert night (setlist + video + photos).
    Reached from the 'concerts' link at the bottom of the dashboard; a Night 1/Night 2
    sub-nav switches between them. No pool data on it."""
    images, videos = scan_concert_dir(concert["dir"], concert["web"])
    web = concert["web"]
    # Lay portrait clips out solo + centered, landscape clips two-up (avoids the tall
    # gap a fixed grid leaves when a vertical clip sits next to a horizontal one).
    # Chronological order is preserved; a landscape isolated next to a portrait just
    # ends up centered on its own row. Orientation read from each mp4's header.
    vid_cells = []
    for v in videos:
        dims = _video_dims(os.path.join(concert["dir"], v))
        orient = "portrait" if dims and dims[1] > dims[0] else "landscape"
        vid_cells.append(
            f'<div class="vid-cell {orient}">'
            f'<video class="concert-video" controls preload="metadata" playsinline>'
            f'<source src="{web}/{v}" type="video/mp4">Your browser can\'t play this video.</video>'
            f'</div>'
        )
    video_html = "".join(vid_cells)
    photo_html = "".join(
        f'<img src="{web}/{f}" loading="lazy" alt="Concert photo">' for f in images
    )
    videos_section = f'<h2 class="concert-sub">Video</h2><div class="concert-videos">{video_html}</div>' if videos else ""
    photos_section = f'<h2 class="concert-sub">Photos</h2><div class="thumbs">{photo_html}</div>' if images else ""
    bits = []
    if images:
        bits.append(f'{len(images)} photo' + ('s' if len(images) != 1 else ''))
    if videos:
        bits.append(f'{len(videos)} video' + ('s' if len(videos) != 1 else ''))
    subtitle = " · ".join(bits)
    links = []
    for c in all_concerts:
        active = ' class="active"' if c["page"] == concert["page"] else ''
        links.append(f'<a href="{c["page"]}"{active}>{esc(c["label"])}</a>')
    night_nav = "".join(links)

    html = f"""<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="utf-8"><meta name="viewport" content="width=device-width, initial-scale=1">
<script src="theme.js?v={build_version}"></script>
<title>{esc(concert['title'])}</title>
<link rel="preconnect" href="https://fonts.googleapis.com">
<link rel="stylesheet" href="https://fonts.googleapis.com/css2?family=DM+Sans:wght@400;500&family=Fraunces:wght@500&display=swap">
<link rel="stylesheet" href="style.css?v={build_version}">
</head>
<body>
<header class="masthead">
<a class="wordmark" href="index.html">Summer 2026 Pool Data</a>
<div style="display:flex;align-items:center;gap:20px;flex-wrap:wrap;justify-content:center;">
<nav><a href="index.html">Dashboard</a><a href="log.html">Data</a><a href="photos.html">Photos</a><a href="Pool_Log.xlsx">Download Full Log</a><a href="Taylor_K2006_Testing_Guide.docx">How I test</a></nav>
<button type="button" id="themeToggle" class="theme-toggle" aria-label="Toggle light/dark theme"></button>
</div>
</header>
<script>
(function () {{
  var btn = document.getElementById('themeToggle');
  function labelToggle() {{
    var next = window.getTheme() === 'dark' ? 'light' : 'dark';
    btn.innerHTML = window.THEME_ICONS[next === 'dark' ? 'moon' : 'sun'];
    btn.setAttribute('aria-label', 'Switch to ' + next + ' mode');
  }}
  labelToggle();
  btn.addEventListener('click', function () {{
    var next = window.getTheme() === 'dark' ? 'light' : 'dark';
    localStorage.setItem('pool_theme', next);
    document.documentElement.setAttribute('data-theme', next);
    location.reload();
  }});
}})();
</script>
<nav class="night-nav">{night_nav}</nav>
<main class="wide">
<h1>{esc(concert['heading'])}</h1>
<p class="cap">{subtitle}. Click a photo to zoom, click again to shrink.</p>
{setlist_html(concert)}
{videos_section}
{photos_section}
</main>
<div class="lightbox" id="lightbox"><img id="lightboxImg" src="" alt=""></div>
<script>
document.querySelectorAll('.thumbs img').forEach(img => {{
  img.addEventListener('click', () => {{
    document.getElementById('lightboxImg').src = img.src;
    document.getElementById('lightbox').classList.add('open');
  }});
}});
document.getElementById('lightbox').addEventListener('click', () => {{
  document.getElementById('lightbox').classList.remove('open');
}});
</script>
</body>
</html>"""
    with open(os.path.join(OUT, concert["page"]), "w", encoding="utf-8") as fh:
        fh.write(html)
    return len(images), len(videos)


def main():
    if os.path.exists(OUT):
        shutil.rmtree(OUT)
    os.makedirs(OUT)

    # Busts browser caching of style.css on every build -- without this, a
    # visitor whose browser cached an older stylesheet could see fresh HTML
    # (new CSS classes) rendered against a stale, non-matching stylesheet.
    build_version = datetime.now().strftime("%Y%m%d%H%M%S")

    tests, doses, all_rows = load_rows()

    shutil.copy(os.path.join(HERE, "site_src", "style.css"), os.path.join(OUT, "style.css"))
    shutil.copy(os.path.join(HERE, "site_src", "theme.js"), os.path.join(OUT, "theme.js"))
    with open(os.path.join(HERE, "site_src", "index.html"), encoding="utf-8") as fh:
        index_html = fh.read()
    index_html = index_html.replace('href="style.css"', f'href="style.css?v={build_version}"')
    index_html = index_html.replace('src="theme.js"', f'src="theme.js?v={build_version}"')
    with open(os.path.join(OUT, "index.html"), "w", encoding="utf-8") as fh:
        fh.write(index_html)

    by_date, videos_by_date = scan_and_resize_photos()
    build_data_json(tests, doses, all_rows, by_date)
    build_log_html(all_rows, set(by_date.keys()) | set(videos_by_date.keys()), build_version)
    build_photos(tests, by_date, videos_by_date, build_version)

    concert_counts = [build_concert_page(c, CONCERTS, build_version) for c in CONCERTS]

    shutil.copy(LOG_FILE, os.path.join(OUT, "Pool_Log.xlsx"))
    if os.path.exists(GUIDE_SRC):
        shutil.copy(GUIDE_SRC, os.path.join(OUT, "Taylor_K2006_Testing_Guide.docx"))

    n_photos = len(os.listdir(os.path.join(OUT, "photos"))) if os.path.isdir(os.path.join(OUT, "photos")) else 0
    concert_summary = ", ".join(f"{c['label']} {ci}+{cv}" for c, (ci, cv) in zip(CONCERTS, concert_counts))
    print(f"Built _site/: {len(tests)} test rows, {len(doses)} dose rows, {n_photos} photos; "
          f"concerts {concert_summary} (photos+videos).")


if __name__ == "__main__":
    main()
