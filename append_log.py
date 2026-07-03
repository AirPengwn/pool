#!/usr/bin/env python3
"""
append_log.py — Append timestamped row(s) to the Log tab of Pool_Log.xlsx IN PLACE.

Edits the existing file so formatting/column widths/styles are preserved. It copies
the cell styling from an existing data row so new rows match.

Column order (fixed):
  Date, Time, Type, pH, FC, CC, FC loss (ppm/24h), Pred next-noon FC,
  Pred error (ppm), TA, CH, CYA, Chlorine added (gal), Cum. Cl (gal),
  Sun (hrs), Rain (in), Fill (gal), Water temp (°F), Weather, Photo, Notes

Row 1 = header, row 2 = SEASON TOTALS (live formulas, see format_log.py) —
new rows always append after the last existing row, so these are untouched.

Derived columns (fill at logging time):
  fc_loss  = prior TEST FC + (gal dosed since)*3.5 - this TEST FC  [on TEST rows]
  pred_fc  = daily_calc.py's "Projected NEXT-NOON FC" (evening model)  [on DOSE rows]
  pred_err = this TEST's FC - the prior DOSE row's pred_fc          [on TEST rows]
  cum_cl   = running season total of chlorine added                [on DOSE rows]
  fill_gal = gallons added via hose (minutes * ~3.1 GPM)      [on EVENT-Fill rows]
  rain_in / sun_hrs from weather.py; photo = dated image filename.

After appending, run `python format_log.py` to re-apply widths, centering, number
formats, alternating-day shading, and the totals-row formulas so new rows match
the rest of the sheet.

Usage (single row):
  python3 append_log.py --date 2026-06-23 --time 12:00 --type TEST \
      --fc 14.0 --cc 0.2 --weather "Sunny, 78F" --notes "25mL x0.2. In band."

Usage (batch from JSON file — preferred when logging several rows at once):
  python3 append_log.py --json rows.json
  where rows.json is a list of objects with any of these keys:
    date, time, type, ph, fc, cc, fc_loss, pred_fc, pred_error, ta, ch, cya,
    chlorine_gal, cum_cl, sun_hrs, rain_in, fill_gal, water_temp, weather,
    photo, notes

Always pass --date and --time and --type for each row. Other fields optional.
"""

import argparse
import json
import os
import sys
from copy import copy
from openpyxl import load_workbook

# Locate the workbook next to this script, so it works from any directory.
LOG_FILE = os.path.join(os.path.dirname(os.path.abspath(__file__)), "Pool_Log.xlsx")
SHEET = "Log"
NCOLS = 22

# JSON/arg key -> column index (1-based)
COLS = {
    "date": 1, "time": 2, "type": 3, "ph": 4, "fc": 5, "cc": 6,
    "fc_loss": 7, "pred_fc": 8, "pred_error": 9, "ta": 10, "ch": 11,
    "cya": 12, "chlorine_gal": 13, "cum_cl": 14, "sun_hrs": 15,
    "rain_in": 16, "fill_gal": 17, "water_temp": 18, "water_level_cm": 19,
    "weather": 20, "photo": 21, "notes": 22,
}


def get_styles(ws, template_row):
    styles = {}
    for c in range(1, NCOLS + 1):
        s = ws.cell(template_row, c)
        styles[c] = (copy(s.font), copy(s.fill), copy(s.alignment),
                     copy(s.border), s.number_format)
    return styles


def write_row(ws, styles, rowdict):
    r = ws.max_row + 1
    for key, col in COLS.items():
        val = rowdict.get(key, None)
        cell = ws.cell(r, col, val)
        f, fl, al, bd, nf = styles[col]
        cell.font = f
        cell.fill = fl
        cell.alignment = al
        cell.border = bd
        cell.number_format = nf
    return r


def main():
    p = argparse.ArgumentParser()
    p.add_argument("--file", default=LOG_FILE)
    p.add_argument("--json", help="Path to JSON file with a list of row objects")
    # single-row args
    p.add_argument("--date"); p.add_argument("--time"); p.add_argument("--type")
    p.add_argument("--ph", type=float); p.add_argument("--fc", type=float)
    p.add_argument("--cc", type=float); p.add_argument("--fc_loss", type=float)
    p.add_argument("--pred_fc", type=float); p.add_argument("--pred_error", type=float)
    p.add_argument("--ta", type=float); p.add_argument("--ch", type=float)
    p.add_argument("--cya")  # may be a string like "~150"
    p.add_argument("--chlorine_gal", type=float)
    p.add_argument("--cum_cl", type=float)
    p.add_argument("--sun_hrs", type=float); p.add_argument("--rain_in", type=float)
    p.add_argument("--fill_gal", type=float)
    p.add_argument("--water_temp", type=float)
    p.add_argument("--weather"); p.add_argument("--photo"); p.add_argument("--notes")
    args = p.parse_args()

    wb = load_workbook(args.file)
    ws = wb[SHEET]
    # Style template = the last existing data row (always already formatted by
    # format_log.py). Dynamic on purpose: row 2 is now a totals row and rows
    # shift on any future insert, so a hardcoded row number breaks silently.
    styles = get_styles(ws, ws.max_row)

    if args.json:
        with open(args.json) as fh:
            rows = json.load(fh)
        if isinstance(rows, dict):
            rows = [rows]
    else:
        if not (args.date and args.time and args.type):
            sys.exit("Single-row mode needs at least --date, --time, --type")
        rows = [{
            "date": args.date, "time": args.time, "type": args.type,
            "ph": args.ph, "fc": args.fc, "cc": args.cc, "fc_loss": args.fc_loss,
            "pred_fc": args.pred_fc, "pred_error": args.pred_error,
            "ta": args.ta, "ch": args.ch, "cya": args.cya,
            "chlorine_gal": args.chlorine_gal, "cum_cl": args.cum_cl,
            "sun_hrs": args.sun_hrs, "rain_in": args.rain_in,
            "fill_gal": args.fill_gal,
            "water_temp": args.water_temp, "weather": args.weather,
            "photo": args.photo, "notes": args.notes,
        }]

    written = [write_row(ws, styles, rd) for rd in rows]
    wb.save(args.file)
    print(f"Appended {len(written)} row(s) at: {written}")


if __name__ == "__main__":
    main()
