#!/usr/bin/env python3
"""
daily_calc.py — one-shot daily calc: weather + FC-loss + dose recommendation.

Reads today's readings from today_input.json (write that file with Write/Edit,
NOT via a Bash argument, so this script's own command line never changes and
stays permanently pre-approved). Prints a full report for Claude to reason
about before writing the final rows.json for append_log.py.

today_input.json fields (fc/cc/date/time required, rest optional):
  {
    "date": "2026-07-01", "time": "12:10",
    "fc": 14.8, "cc": 0.2,
    "ph": null, "ta": null, "ch": null, "cya": null,
    "water_temp": 81.0,
    "weather_override": "sunny",   // use when auto bucket misjudges the day
    "rain_in_override": null       // use when auto daily precip is misleading
  }

If the last DOSE row has a 'Pred next-noon FC' value, this also auto-scores
it against today's actual FC and prints the 'Pred error (ppm)' to log.

Usage:  python daily_calc.py   (no args — reads today_input.json next to it)
"""

import json
import os
import sys

HERE = os.path.dirname(os.path.abspath(__file__))
sys.path.insert(0, HERE)

import weather as W
import dose as D
from append_log import COLS as LOGCOLS, LOG_FILE, SHEET

INPUT_FILE = os.path.join(HERE, "today_input.json")


def last_test_and_doses_since(ws):
    """Return (last_test_fc, last_test_row, gallons_dosed_since, last_pred_fc,
    last_pred_row) by scanning from the bottom for the most recent TEST row
    with an FC value, then summing Chlorine added (gal) on every row after it
    and noting the most recent DOSE row's 'Pred next-noon FC' in that same
    window (last night's prediction for today — used to auto-score it)."""
    last_row = None
    for r in range(ws.max_row, 2, -1):
        if ws.cell(r, LOGCOLS["type"]).value == "TEST" and \
                isinstance(ws.cell(r, LOGCOLS["fc"]).value, (int, float)):
            last_row = r
            break
    if last_row is None:
        return None, None, 0.0, None, None
    doses = 0.0
    last_pred_fc, last_pred_row = None, None
    for r in range(last_row + 1, ws.max_row + 1):
        v = ws.cell(r, LOGCOLS["chlorine_gal"]).value
        if isinstance(v, (int, float)):
            doses += v
        pf = ws.cell(r, LOGCOLS["pred_fc"]).value
        if isinstance(pf, (int, float)):
            last_pred_fc, last_pred_row = pf, r
    return ws.cell(last_row, LOGCOLS["fc"]).value, last_row, doses, last_pred_fc, last_pred_row


def main():
    if not os.path.exists(INPUT_FILE):
        sys.exit(f"missing {INPUT_FILE} — write today's readings there first")
    with open(INPUT_FILE) as fh:
        inp = json.load(fh)

    # --- weather ---
    try:
        wd = W.fetch(W.CONFIG)
        cur, daily = wd["current"], wd["daily"]
        dsf = W.day_so_far(cur, wd.get("hourly"))
        auto_bucket = W.classify(cur, daily, W.CONFIG, dsf)
        weather_str = W.log_string(cur, daily, dsf)
    except Exception as e:
        cur = daily = dsf = None
        auto_bucket = "partly"
        weather_str = f"(weather fetch failed: {e})"

    bucket = inp.get("weather_override") or auto_bucket
    sun_hrs_today = None
    rain_in_today = inp.get("rain_in_override")
    if daily is not None:
        rain_in_today = rain_in_today if rain_in_today is not None else (daily["precipitation_sum"][0] or 0.0)
    try:
        hist = W.fetch_history(W.CONFIG, 0)
        if hist:
            sun_hrs_today = hist[-1]["sun_hrs"]
    except Exception:
        pass

    # --- FC loss vs last TEST ---
    from openpyxl import load_workbook
    wb = load_workbook(LOG_FILE)
    ws = wb[SHEET]
    last_fc, last_row, doses_since, last_pred_fc, last_pred_row = last_test_and_doses_since(ws)
    fc_loss = None
    if last_fc is not None:
        fc_loss = round(last_fc + doses_since * D.CONFIG["ppm_per_gallon"] - inp["fc"], 1)

    # --- auto-score last night's prediction against today's actual FC ---
    pred_error = None
    if last_pred_fc is not None:
        pred_error = round(inp["fc"] - last_pred_fc, 1)
        if pred_error == 0:
            pred_error = 0.0  # normalize -0.0

    # --- dose recommendation ---
    dose_text = D.calc(inp["fc"], inp.get("cc", 0.0), bucket, D.CONFIG["cya_current"], D.CONFIG)

    print("=" * 60)
    print(f"DATE {inp['date']}  TIME {inp['time']}")
    print(f"Weather (auto)   : {weather_str}")
    print(f"Auto bucket      : {auto_bucket}   | Using bucket: {bucket}"
          + ("  (OVERRIDDEN)" if inp.get("weather_override") else ""))
    if sun_hrs_today is not None:
        print(f"Sun today (full-day est.): {sun_hrs_today} hrs")
    if rain_in_today is not None:
        print(f"Rain today       : {rain_in_today} in")
    print(f"Last TEST        : row {last_row}, FC {last_fc}"
          if last_row else "Last TEST        : none found")
    print(f"Gallons dosed since last TEST: {doses_since}")
    print(f"Implied FC loss  : {fc_loss} ppm" if fc_loss is not None else "Implied FC loss  : n/a")
    if last_pred_fc is not None:
        print("-" * 60)
        print("PREDICTION CHECK (forward validation):")
        print(f"  Last evening's prediction (row {last_pred_row}): {last_pred_fc}")
        print(f"  Today's actual FC: {inp['fc']}")
        print(f"  Pred error (ppm): {pred_error:+.1f}"
              "   -> log this as today's TEST row 'Pred error (ppm)'")
    print("-" * 60)
    print(dose_text)
    print("-" * 60)
    # --- EVENING-DOSE PLAN (calibrated L24 model, 2026-07-01) ---
    plan = D.evening_plan(inp["fc"], sun_hrs_today, inp.get("water_temp"), D.CONFIG)
    print("EVENING-DOSE PLAN (calibrated):")
    print(f"  L24 (24h loss est.): {plan['l24']} ppm"
          f"   [sun {sun_hrs_today}h, water {inp.get('water_temp')}F]")
    print(f"  Recommended EVENING dose: {plan['dose']:.1f} gal (raw {plan['raw_gal']:.2f})"
          f"  -> targets aim {plan['aim']} at next noon")
    print(f"  Projected NEXT-NOON FC: ~{plan['projected_next_noon']:.1f}"
          "  -> log as 'Pred next-noon FC' on tonight's DOSE row")
    print("  (If your final pour differs from the recommended dose above, recompute with:")
    print(f"   python dose.py --evening --fc {inp['fc']} --sun {sun_hrs_today} "
          f"--water_temp {inp.get('water_temp')} --dose FINAL_GAL)")
    print("=" * 60)


if __name__ == "__main__":
    main()
