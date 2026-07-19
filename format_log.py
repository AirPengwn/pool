#!/usr/bin/env python3
"""
format_log.py — (re)apply consistent formatting to the Log tab. Idempotent.

Run this AFTER append_log.py so newly added rows match the rest of the sheet.
It does NOT touch data — only presentation:
  - removes the header auto-filter
  - sets sensible column widths
  - centers all cells EXCEPT Weather & Notes (left-aligned, word-wrapped)
  - shades alternating DATES light gray (all rows of one date share a shade;
    the next date is unshaded; etc.)
  - resets row heights to auto (Excel right-sizes on open)

Column layout (1-based):
  1 Date  2 Time  3 Type  4 pH  5 FC  6 CC  7 FC loss (ppm/24h)
  8 Pred next-noon FC  9 Pred error (ppm)  10 TA  11 CH  12 CYA
  13 Chlorine added (gal)  14 Cum. Cl (gal)  15 Sun (hrs)  16 Rain (in)
  17 Fill (gal)  18 Water temp (F)  19 Water level (cm)  20 Weather
  21 Photo  22 Notes  23 Load

Row 1 = header. Row 2 = SEASON TOTALS (live SUM formulas over Chlorine/Rain/
Fill columns; A2:D2 merged for the label). Data starts row 3. Top two rows
are frozen (freeze_panes = A3).

Usage:  python format_log.py
"""

import os
from copy import copy
from openpyxl import load_workbook
from openpyxl.styles import Alignment, PatternFill, Border, Side, Font
from openpyxl.utils import get_column_letter

# Locate the workbook next to this script, so it works from any directory.
FILE = os.path.join(os.path.dirname(os.path.abspath(__file__)), "Pool_Log.xlsx")
SHEET = "Log"
NCOLS = 23
DATA_START_ROW = 3   # row 1 = header, row 2 = totals, data from row 3
# Column widths — captured from John's manual Excel adjustments (2026-06-27),
# with Fill (gal) inserted at 15 on 2026-06-30, Pred next-noon FC / Pred error
# inserted at 8/9 (shifting TA..Notes right by 2) on 2026-07-01, Water level (cm)
# inserted at 19 (shifting Weather/Photo/Notes right by 1) on 2026-07-02. Edit here to change.
WIDTHS = {1: 15.14, 2: 17.71, 3: 15.71, 4: 6.0, 5: 13.0, 6: 13.0, 7: 12.14,
          8: 13.0, 9: 12.0, 10: 6.0, 11: 13.0, 12: 8.0, 13: 12.0, 14: 11.0,
          15: 9.0, 16: 8.0, 17: 10.0, 18: 11.0, 19: 13.0, 20: 38.0, 21: 16.0, 22: 83.86,
          23: 14.0}   # Load added 2026-07-19
MIN_ROW_HEIGHT = 20  # floor so single-line rows aren't crowded; tall rows auto-grow
LEFT_WRAP = {20, 22}           # Weather, Notes -> left + wrap
NUMFMT = {4: "0.0", 5: "0.0", 6: "0.0", 7: "0.0", 8: "0.0",
          9: "+0.0;-0.0;0.0",  # signed: shows explicit + / - so over- vs under-prediction reads at a glance
          10: "0", 11: "0", 12: "0",
          13: "0.0", 14: "0.0", 15: "0.0", 16: "0.00", 17: "0.0",
          18: "0.0", 19: "0.0"}  # 18=water temp, 19=water level: 1 decimal, never round
# One consistent font everywhere (was a Calibri/Arial mix).
HEADER_FONT = Font(name="Calibri", size=11, bold=True, color="FFFFFFFF")
DATA_FONT = Font(name="Calibri", size=11)
# Row 2 = running-totals summary row (frozen under the header).
TOTALS_FONT = Font(name="Calibri", size=11, bold=True)
# IMPORTANT: fill colors must be 8-digit ARGB (leading FF = opaque). A 6-digit
# hex makes openpyxl default the alpha byte to "00" (fully transparent), which
# renders WHITE in Excel — that bug had turned the totals row AND the gray
# day-shading invisible (2026-07-01).
HEADER_FILL = PatternFill("solid", fgColor="FF1F6FB2")   # row 1 blue (hardcoded so it can't silently degrade to white)
TOTALS_FILL = PatternFill("solid", fgColor="FF9DC3E6")   # row 2 blue — stands out under the header
GRAY = PatternFill("solid", fgColor="FFC4C4C4")          # alternating-day shading
NOFILL = PatternFill(fill_type=None)

# Thin grid on every cell (dark enough to show on the gray shade);
# thick dark box around each day block.
THIN = Side(style="thin", color="808080")
THICK = Side(style="medium", color="333333")


def main():
    wb = load_workbook(FILE)
    ws = wb[SHEET]

    ws.auto_filter.ref = None   # remove header filtering

    for c, w in WIDTHS.items():
        ws.column_dimensions[get_column_letter(c)].width = w

    prev_date, shade = None, True   # first distinct date ends up unshaded
    for r in range(1, ws.max_row + 1):
        if r == 1:  # header: uniform bold white Calibri on the blue fill
            for c in range(1, NCOLS + 1):
                hc = ws.cell(r, c)
                hc.font = copy(HEADER_FONT)
                hc.fill = copy(HEADER_FILL)
                hc.alignment = Alignment(
                    horizontal="center", vertical="center", wrap_text=True)
                hc.border = Border(
                    left=THICK if c == 1 else THIN,
                    right=THICK if c == NCOLS else THIN,
                    top=THICK, bottom=THICK)
            ws.row_dimensions[r].height = 28
            continue

        if r == 2:  # SEASON TOTALS row: bold, light-blue, thick box, own number formats
            for c in range(1, NCOLS + 1):
                tc = ws.cell(r, c)
                tc.font = copy(TOTALS_FONT)
                tc.fill = copy(TOTALS_FILL)
                tc.alignment = Alignment(
                    horizontal="center", vertical="center", wrap_text=True)
                if c in NUMFMT:
                    tc.number_format = NUMFMT[c]
                tc.border = Border(
                    left=THICK if c == 1 else THIN,
                    right=THICK if c == NCOLS else THIN,
                    top=THICK, bottom=THICK)
            ws.cell(2, 1).alignment = Alignment(
                horizontal="left", vertical="center")
            ws.row_dimensions[r].height = 20
            continue

        d = ws.cell(r, 1).value
        new_date = (d != prev_date)
        if new_date:
            shade = not shade       # toggle once per distinct date
            prev_date = d
        # last row of this day's block? (next row is a different date / sheet end)
        last_of_day = (r == ws.max_row) or (ws.cell(r + 1, 1).value != d)
        fill = GRAY if shade else NOFILL

        for c in range(1, NCOLS + 1):
            cell = ws.cell(r, c)
            cell.font = copy(DATA_FONT)
            if c in LEFT_WRAP:
                cell.alignment = Alignment(
                    horizontal="left", vertical="center", wrap_text=True)
            else:
                cell.alignment = Alignment(
                    horizontal="center", vertical="center", wrap_text=False)
            cell.fill = fill
            if c in NUMFMT:
                cell.number_format = NUMFMT[c]
            # thin grid everywhere; thick on the day-block's outer edges
            cell.border = Border(
                left=THICK if c == 1 else THIN,
                right=THICK if c == NCOLS else THIN,
                top=THICK if new_date else THIN,
                bottom=THICK if last_of_day else THIN)
        # Row height: if the wrapped columns (Weather/Notes) all fit on one line,
        # this is a short row -> pin to MIN_ROW_HEIGHT so it isn't crowded.
        # If any wrap to multiple lines, leave auto (None) so Excel grows it.
        multiline = any(
            ws.cell(r, c).value is not None
            and len(str(ws.cell(r, c).value)) > WIDTHS.get(c, 999)
            for c in LEFT_WRAP)
        ws.row_dimensions[r].height = None if multiline else MIN_ROW_HEIGHT

    ws.freeze_panes = "A3"   # lock header (row 1) + totals row (row 2)
    wb.save(FILE)
    print("Formatted Log: filter removed, widths/alignment set, "
          "alternating-day shading applied, totals row styled, "
          "top 2 rows frozen, row heights auto.")


if __name__ == "__main__":
    main()
