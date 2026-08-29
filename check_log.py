#!/usr/bin/env python3
"""check_log.py -- assert the invariants of Pool_Log.xlsx. Exit 1 on any ERROR.

WHY THIS EXISTS (2026-08-29). Three separate silent-write failures surfaced in
one session, all the same shape: something accepted bad input or failed to
apply, said nothing, and the damage sat in the log undetected.

  1. POOL.md edits via str.replace() whose anchor did not match byte-for-byte.
     The script printed "updated" and changed nothing. Four rules lost.
  2. A row written with "water_level" instead of "water_level_cm". append_log
     iterates its known columns and pulls from the dict, so the misspelled key
     was discarded. The cell landed empty -- and an empty level silently
     corrupts the NEXT day's mass balance, because analyze_loss.py carries the
     last known level forward.
  3. Photo cells holding a range shorthand ("2026-06-27_1-3.jpg") that names no
     real file. It renders as plain text in the log table, so it never visibly
     broke; it survived two months.

The common thread is NOT carelessness, it is that nothing ever read the data
back. A success message is not evidence. This script is the read-back.

Run it after every logging session, before committing. It also runs from the
pre-commit hook (tools/pre-commit) and from the deploy workflow, so a log that
violates an invariant cannot reach the site.

Usage:  python check_log.py            # ERRORs exit 1, WARNs do not
        python check_log.py --strict   # WARNs also exit 1
"""

import glob
import os
import re
import sys
from datetime import date, timedelta

import openpyxl

HERE = os.path.dirname(os.path.abspath(__file__))
LOG = os.path.join(HERE, "Pool_Log.xlsx")
PHOTODIR = os.path.join(HERE, "photos")

HEADER = ["Date", "Time", "Type", "pH", "FC", "CC", "FC loss (ppm/24h)",
          "Pred next-noon FC", "Pred error (ppm)", "TA", "CH", "CYA",
          "Chlorine added (gal)", "Cum. Cl (gal)", "Sun (hrs)", "Rain (in)",
          "Fill (gal)", "Water temp (°F)", "Water level (cm)", "Weather",
          "Photo", "Notes", "Load"]

# Mass-balance constants -- must match analyze_loss.py exactly.
FULL_GAL, FULL_LEVEL_CM, GAL_PER_CM = 31400.0, 15.81, 196.3
CL_MASS_PER_GAL = 3.5 * 34400.0
GOOD_FROM = date(2026, 6, 24)          # fresh FAS-DPD onwards

# Load column: canonical[+canonical]; optional human qualifier.
# POOL.md is the source of truth -- keep the two in sync.
VOCAB = {"quiet", "swimmers", "party", "dog", "storm", "smoke", "millipedes",
         "caterpillars", "leaves", "earthworms", "pillbugs", "carcass",
         "burn debris", "frogs", "insects", "seeds"}

# Fields are only MANDATORY from the date each became standard practice, and
# only on the first (noon) test of a day -- a repeat test the same day does not
# need a fresh level or photo. Cutoff is deliberately later than the last
# historical gap so the checker guards TODAY without re-litigating June.
REQUIRED_FROM = date(2026, 8, 1)

ERRORS, WARNS = [], []


def err(check, msg):
    ERRORS.append((check, msg))


def warn(check, msg):
    WARNS.append((check, msg))


def as_date(v):
    if hasattr(v, "year"):
        return date(v.year, v.month, v.day)
    if isinstance(v, str) and "-" in v:
        try:
            y, m, d = v.split("-")[:3]
            return date(int(y), int(m), int(d))
        except ValueError:
            return None
    return None


def num(v):
    return v if isinstance(v, (int, float)) else None


def main():
    strict = "--strict" in sys.argv
    if not os.path.exists(LOG):
        sys.exit(f"missing {LOG}")
    ws = openpyxl.load_workbook(LOG, data_only=True)["Log"]

    # ---- 1. schema ---------------------------------------------------------
    hdr = [ws.cell(1, c).value for c in range(1, len(HEADER) + 1)]
    if hdr != HEADER:
        for i, (a, b) in enumerate(zip(HEADER, hdr), start=1):
            if a != b:
                err("schema", f"col {i}: expected {a!r}, found {b!r}")
    if ws.max_column != len(HEADER):
        err("schema", f"expected {len(HEADER)} columns, sheet has {ws.max_column}")

    rows = [r for r in range(3, ws.max_row + 1) if ws.cell(r, 1).value]
    g = lambda r, c: ws.cell(r, c).value

    # ---- 2. row basics -----------------------------------------------------
    tests, doses = [], []
    for r in rows:
        d, t = as_date(g(r, 1)), str(g(r, 3) or "")
        if d is None:
            err("dates", f"row {r}: unparseable date {g(r, 1)!r}")
            continue
        if not t:
            err("types", f"row {r}: no Type")
        if t == "TEST":
            tests.append((r, d))
        elif t == "DOSE":
            doses.append((r, d))
            if num(g(r, 13)) is None:
                err("doses", f"row {r} {d}: DOSE row with no gallons")
        if t != "DOSE" and num(g(r, 13)) is not None:
            err("doses", f"row {r} {d}: chlorine logged on a {t!r} row, not DOSE")

    # first (earliest-time) test of each day; repeat tests are exempt below
    byday = {}
    for r, d in tests:
        byday.setdefault(d, []).append(r)
    first_test = {d: min(rs) for d, rs in byday.items()}

    # ---- 3. calendar continuity -------------------------------------------
    if byday:
        lo, hi = min(byday), max(byday)
        missing = [(lo + timedelta(days=i)).isoformat()
                   for i in range((hi - lo).days + 1)
                   if (lo + timedelta(days=i)) not in byday]
        if missing:
            err("calendar", f"{len(missing)} day(s) between {lo} and {hi} "
                            f"have no TEST row: {missing[:8]}")

    # ---- 4. required fields on the day's first test ------------------------
    for d, r in sorted(first_test.items()):
        if d < REQUIRED_FROM:
            continue
        for col, name in ((5, "FC"), (19, "water level"), (21, "photo"),
                          (23, "load"), (2, "time")):
            v = g(r, col)
            missing = (num(v) is None) if col in (5, 19) else not v
            if missing:
                err("required", f"row {r} {d}: first test of the day has no {name}")

    # ---- 5. derived: Cum. Cl is the running total --------------------------
    run = 0.0
    for r in rows:
        v = num(g(r, 13))
        if v is not None:
            run += v
        c = num(g(r, 14))
        if c is not None and abs(c - run) > 1e-6:
            err("cum_cl", f"row {r}: Cum. Cl {c} != running total {round(run, 2)}")

    # ---- 6. derived: FC loss is the mass balance ---------------------------
    tv = {}
    for r, d in tests:
        if d not in tv:
            tv[d] = dict(fc=num(g(r, 5)), level=num(g(r, 19)), row=r,
                         loss=num(g(r, 7)))
    dose_by_day = {}
    for r, d in doses:
        dose_by_day[d] = dose_by_day.get(d, 0.0) + (num(g(r, 13)) or 0.0)
    last = None
    for d in sorted(tv):
        if tv[d]["level"] is not None:
            last = tv[d]["level"]
        else:
            tv[d]["level"] = last
    vol = lambda cm: FULL_GAL + (cm - FULL_LEVEL_CM) * GAL_PER_CM
    for d in sorted(tv):
        p, b = tv.get(d - timedelta(days=1)), tv[d]
        if not p or (d - timedelta(days=1)) < GOOD_FROM:
            continue
        if None in (p["fc"], b["fc"], p["level"], b["level"]):
            continue
        v0, v1 = vol(p["level"]), vol(b["level"])
        mb = round((v0 * p["fc"] + dose_by_day.get(d - timedelta(days=1), 0.0)
                    * CL_MASS_PER_GAL - v1 * b["fc"]) / ((v0 + v1) / 2), 1)
        if b["loss"] is None:
            err("fc_loss", f"row {b['row']} {d}: FC loss blank, should be {mb}")
        elif abs(b["loss"] - mb) > 0.05:
            err("fc_loss", f"row {b['row']} {d}: FC loss {b['loss']} != "
                           f"mass balance {mb} -- do not hand-compute it")

    # ---- 7. photo referential integrity, both directions -------------------
    referenced, seen_twice = {}, []
    for r in rows:
        p = g(r, 21)
        if not p:
            continue
        for f in [x.strip() for x in str(p).split(";") if x.strip()]:
            if not re.fullmatch(r"\d{4}-\d{2}-\d{2}_(?:v)?\d+\.(?:jpe?g|png|mp4)", f, re.I):
                err("photos", f"row {r}: {f!r} is not a dated filename "
                              f"(range shorthand like _1-3.jpg is NOT a file)")
            if f in referenced:
                seen_twice.append((f, referenced[f], r))
            referenced[f] = r
    for f, a, b in seen_twice:
        err("photos", f"{f} referenced by both row {a} and row {b}")
    on_disk = {os.path.basename(x) for x in glob.glob(os.path.join(PHOTODIR, "*"))
               if re.search(r"\.(jpe?g|png)$", x, re.I)}
    for f in sorted(set(referenced) - on_disk):
        if not f.lower().endswith(".mp4"):
            err("photos", f"row {referenced[f]}: {f} referenced but not in photos/")
    for f in sorted(on_disk - set(referenced)):
        if re.match(r"^\d{4}-\d{2}-\d{2}_", f):
            err("photos", f"{f} is on disk but referenced by no row")
        else:
            warn("photos", f"{f} is an unfiled camera original in photos/")

    # ---- 8. Load vocabulary ------------------------------------------------
    for r in rows:
        v = g(r, 23)
        if not v:
            continue
        for tok in [t.strip() for t in str(v).split(";")[0].split("+")]:
            if tok not in VOCAB:
                err("load", f"row {r}: load tag {tok!r} is not in the "
                            f"controlled vocabulary (POOL.md)")

    # ---- report ------------------------------------------------------------
    checks = ["schema", "dates", "types", "doses", "calendar", "required",
              "cum_cl", "fc_loss", "photos", "load"]
    failed = {c for c, _ in ERRORS}
    print(f"check_log.py -- {len(rows)} data rows, {len(byday)} tested days\n")
    for c in checks:
        n = sum(1 for k, _ in ERRORS if k == c)
        print(f"  {'FAIL' if c in failed else 'ok  '}  {c:<10}"
              + (f"  {n} problem(s)" if n else ""))
    if ERRORS:
        print(f"\n{len(ERRORS)} ERROR(S):")
        for c, m in ERRORS:
            print(f"  [{c}] {m}")
    if WARNS:
        print(f"\n{len(WARNS)} warning(s):")
        for c, m in WARNS:
            print(f"  [{c}] {m}")
    if not ERRORS and not WARNS:
        print("\nall invariants hold")
    sys.exit(1 if ERRORS or (strict and WARNS) else 0)


if __name__ == "__main__":
    main()
